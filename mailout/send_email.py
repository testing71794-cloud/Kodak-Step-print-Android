"""
Send final_execution_report.xlsx after parallel orchestration completes (HTML + attachments).

Default attachments (Gmail-safe, max 18MB per file / 20MB total):
  1) final_execution_report.xlsx
  2) failed_tests_artifacts.zip when present and under size limits; otherwise Jenkins artifact link
  3) Optional AI files when ORCH_EMAIL_ATTACH_AI=1

Email body lists failed tests only (FAIL/FLAKY/PARSE_ERROR/ERROR) with build summary,
failure reasons, AI analysis, and screenshot/video links when archived.
"""
from __future__ import annotations

import html
import json
import logging
import os
import socket
import smtplib
import ssl
import sys
import time
import zipfile
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import NamedTuple

from openpyxl import load_workbook

_REPO = Path(__file__).resolve().parents[1]
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))
from utils.device_utils import render_device_display  # noqa: E402
from utils.git_branch import detect_git_branch  # noqa: E402
from utils.project_identity import EXECUTION_SUMMARY_TITLE, PROJECT_DISPLAY_NAME  # noqa: E402

logger = logging.getLogger("orch.mail")

MAX_ATTACHMENT_SIZE = 18 * 1024 * 1024  # 18MB per file (Gmail-safe)
MAX_TOTAL_ATTACHMENT_SIZE = 20 * 1024 * 1024  # 20MB total before link-only mode
EMAIL_FAILURE_STATUSES = frozenset({"FAIL", "FLAKY", "PARSE_ERROR", "ERROR"})
SIZE_SKIP_MSG = "Attachment skipped because it exceeded Gmail size limits."


def resolve_final_excel_path(root: Path) -> Path | None:
    """
    Locate final_execution_report.xlsx for attachment.
    Order: explicit env path → repo root → build-summary → shallowest rglob under root.
    """
    root = root.resolve()
    for env_name in ("FINAL_EXECUTION_REPORT_XLSX", "ORCH_EXCEL_OUT"):
        raw = os.getenv(env_name, "").strip()
        if not raw:
            continue
        p = Path(raw)
        if not p.is_absolute():
            p = (root / p).resolve()
        if p.is_file():
            logger.info("Using Excel from %s=%s", env_name, p)
            return p
        logger.warning("Env %s points to missing file: %s", env_name, p)

    candidates = [
        root / "final_execution_report.xlsx",
        root / "build-summary" / "final_execution_report.xlsx",
    ]
    for c in candidates:
        if c.is_file():
            logger.info("Using Excel: %s", c)
            return c

    matches = [p for p in root.rglob("final_execution_report.xlsx") if p.is_file()]
    if matches:
        best = min(matches, key=lambda p: (len(p.parts), str(p)))
        logger.info("Using Excel (search): %s", best)
        return best

    logger.error(
        "No final_execution_report.xlsx under %s (tried env FINAL_EXECUTION_REPORT_XLSX / ORCH_EXCEL_OUT, root, build-summary, rglob)",
        root,
    )
    return None


def resolve_execution_logs_zip(excel_path: Path, root: Path | None) -> Path | None:
    """Use existing build-summary/execution_logs.zip (or next to the Excel) when present."""
    candidates: list[Path] = [excel_path.parent / "execution_logs.zip"]
    if root is not None:
        r = root.resolve()
        candidates.extend(
            [
                r / "build-summary" / "execution_logs.zip",
                r / "execution_logs.zip",
            ]
        )
    seen: set[Path] = set()
    for p in candidates:
        rp = p.resolve()
        if rp in seen:
            continue
        seen.add(rp)
        if rp.is_file():
            logger.info("Found execution logs zip: %s", rp)
            return rp
    return None


def _collect_log_files_for_zip(root: Path) -> list[Path]:
    """
    All *.log under reports/ and status/ (orchestrator + Maestro logs) for execution_logs.zip.
    """
    r = root.resolve()
    seen: set[Path] = set()
    out: list[Path] = []
    for sub in ("reports", "status", "collected-artifacts"):
        d = r / sub
        if not d.is_dir():
            continue
        for p in d.rglob("*.log"):
            if p.is_file():
                k = p.resolve()
                if k not in seen:
                    seen.add(k)
                    out.append(p)
    return sorted(out, key=lambda p: str(p))


def build_execution_logs_zip(root: Path) -> Path | None:
    """
    Create build-summary/execution_logs.zip from reports/**/*.log when any logs exist.
    """
    r = root.resolve()
    log_files = _collect_log_files_for_zip(r)
    if not log_files:
        return None
    out_dir = r / "build-summary"
    out_dir.mkdir(parents=True, exist_ok=True)
    zpath = out_dir / "execution_logs.zip"
    with zipfile.ZipFile(zpath, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in log_files:
            try:
                arc = p.relative_to(r)
            except ValueError:
                arc = p.name
            zf.write(p, arcname=str(arc).replace("\\", "/"))
    logger.info("Created %s with %d log file(s)", zpath, len(log_files))
    return zpath


def resolve_or_build_execution_logs_zip(excel_path: Path, root: Path | None) -> Path | None:
    """Prefer an existing execution_logs.zip; otherwise zip all reports/**/*.log if any."""
    z = resolve_execution_logs_zip(excel_path, root)
    if z is not None:
        return z
    if root is None:
        return None
    return build_execution_logs_zip(root)


def load_failed_tests_summary(root: Path) -> tuple[list[dict], bool]:
    """
    Read build-summary/failed_tests_summary.json written by collect_failed_artifacts.py.
    Returns (failure rows, True) when the summary file exists; otherwise ([], False).
    """
    path = root.resolve() / "build-summary" / "failed_tests_summary.json"
    if not path.is_file():
        return [], False
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning("Could not read failed_tests_summary.json: %s", exc)
        return [], True
    rows = data.get("failures") if isinstance(data, dict) else None
    if not isinstance(rows, list):
        return [], True
    return [r for r in rows if isinstance(r, dict)], True


def resolve_failed_tests_artifacts_zip(root: Path) -> Path | None:
    p = root.resolve() / "build-summary" / "failed_tests_artifacts.zip"
    return p if p.is_file() else None


def _file_size(path: Path) -> int:
    try:
        return path.stat().st_size
    except OSError:
        return 0


def _format_bytes(num: int) -> str:
    if num < 1024:
        return f"{num} B"
    if num < 1024 * 1024:
        return f"{num / 1024:.1f} KB"
    return f"{num / (1024 * 1024):.2f} MB"


def _jenkins_build_url() -> str:
    raw = getenv_any("BUILD_URL", "JENKINS_BUILD_URL", default="").rstrip("/")
    return raw


def _jenkins_artifact_url(relative_path: str) -> str | None:
    base = _jenkins_build_url()
    if not base:
        return None
    rel = relative_path.lstrip("/").replace("\\", "/")
    return f"{base}/artifact/{rel}"


def _is_failure_status(status: str) -> bool:
    return (status or "").upper() in EMAIL_FAILURE_STATUSES


def _filter_failed_email_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [r for r in rows if _is_failure_status(r.get("status", ""))]


def _build_email_subject(*, failed_count: int, total_count: int) -> str:
    build_no = getenv_any("BUILD_NUMBER", "BUILD_ID", default="").strip()
    build_suffix = f" | Build #{build_no}" if build_no else ""
    if failed_count > 0:
        return f"[FAIL] {PROJECT_DISPLAY_NAME} | {failed_count} Failed Tests{build_suffix}"
    return f"[PASS] {PROJECT_DISPLAY_NAME}{build_suffix}"


def _compute_pass_rate(passed: int, total: int) -> str:
    if total <= 0:
        return "0.00%"
    return f"{(passed / total) * 100:.2f}%"


def _enrich_rows_from_failed_summary(
    rows: list[dict[str, str]], failed_summary: list[dict]
) -> list[dict[str, str]]:
    if not failed_summary:
        return rows
    by_key: dict[tuple[str, str, str], dict] = {}
    for item in failed_summary:
        suite = str(item.get("suite") or "").strip().casefold()
        flow = str(item.get("test_name") or item.get("flow") or "").strip()
        dev = str(item.get("device_id") or "").strip()
        by_key[(suite, flow, dev)] = item

    out: list[dict[str, str]] = []
    for row in rows:
        nr = dict(row)
        key = (
            str(nr.get("suite") or "").strip().casefold(),
            str(nr.get("flow") or "").strip(),
            str(nr.get("device_id") or "").strip(),
        )
        hit = by_key.get(key)
        if hit:
            if not nr.get("failure_reason") or nr.get("failure_reason") == "—":
                nr["failure_reason"] = str(
                    hit.get("failure_reason") or hit.get("reason") or "—"
                )
            nr["video_artifact"] = str(hit.get("video_artifact") or "")
            nr["screenshot_artifact"] = str(hit.get("screenshot_artifact") or "")
        out.append(nr)
    return out


class _AttachmentPlan(NamedTuple):
    attach_paths: list[Path]
    attach_labels: list[str]
    download_links: list[tuple[str, str]]
    warnings: list[str]
    link_only_mode: bool
    total_bytes: int


def _plan_attachments(
    candidates: list[tuple[Path, str]],
    *,
    force_link_names: frozenset[str] | None = None,
) -> _AttachmentPlan:
    attach_paths: list[Path] = []
    attach_labels: list[str] = []
    download_links: list[tuple[str, str]] = []
    warnings: list[str] = []
    total = 0
    force_link_names = force_link_names or frozenset()
    link_only_mode = False

    for path, label in candidates:
        if not path.is_file():
            continue
        size = _file_size(path)
        rel = None
        try:
            rel = str(path.relative_to(_REPO)).replace("\\", "/")
        except ValueError:
            rel = f"build-summary/{path.name}"

        url = _jenkins_artifact_url(rel) if rel else None
        if url:
            download_links.append((label, url))

        force_link = path.name in force_link_names or size > MAX_ATTACHMENT_SIZE
        if force_link:
            warnings.append(
                f"{label} ({_format_bytes(size)}): "
                + (SIZE_SKIP_MSG if size > MAX_ATTACHMENT_SIZE else "sent as Jenkins download link.")
            )
            if size > MAX_ATTACHMENT_SIZE:
                logger.warning(
                    "[gcp-email] skipped attachment %s size=%s",
                    path.name,
                    _format_bytes(size),
                )
            link_only_mode = True
            continue

        if total + size > MAX_TOTAL_ATTACHMENT_SIZE:
            warnings.append(
                f"{label} ({_format_bytes(size)}): total attachment size would exceed "
                f"{_format_bytes(MAX_TOTAL_ATTACHMENT_SIZE)}; using download link."
            )
            link_only_mode = True
            continue

        attach_paths.append(path)
        attach_labels.append(f"{label} ({_format_bytes(size)})")
        total += size

    return _AttachmentPlan(
        attach_paths=attach_paths,
        attach_labels=attach_labels,
        download_links=download_links,
        warnings=warnings,
        link_only_mode=link_only_mode,
        total_bytes=total,
    )


def _add_file_attachment(msg: EmailMessage, path: Path) -> None:
    data = path.read_bytes()
    name = path.name
    ext = path.suffix.lower()
    if ext == ".json":
        msg.add_attachment(data, maintype="application", subtype="json", filename=name)
    elif ext == ".xlsx":
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=name,
        )
    elif ext == ".zip":
        msg.add_attachment(data, maintype="application", subtype="zip", filename=name)
    else:
        msg.add_attachment(data, maintype="application", subtype="octet-stream", filename=name)
    logger.info("Attached: %s (%s)", name, _format_bytes(len(data)))


def _truthy_env(name: str, default: str = "") -> bool:
    return getenv_any(name, default=default).lower() in ("1", "true", "yes", "on")


def _smtp_ssl_context() -> ssl.SSLContext:
    if _truthy_env("SMTP_SSL_VERIFY", default="1"):
        return ssl.create_default_context()
    logger.warning("[gcp-email] SMTP_SSL_VERIFY=0 — TLS certificate verification disabled")
    return ssl._create_unverified_context()


def _format_smtp_error(exc: BaseException) -> str:
    parts = [f"{type(exc).__name__}: {exc}"]
    if isinstance(exc, smtplib.SMTPResponseException):
        parts.append(f"smtp_code={exc.smtp_code}")
        try:
            parts.append(f"smtp_error={exc.smtp_error!r}")
        except Exception:
            pass
    if isinstance(exc, OSError):
        if getattr(exc, "errno", None) is not None:
            parts.append(f"errno={exc.errno}")
        if getattr(exc, "strerror", None):
            parts.append(f"strerror={exc.strerror}")
        if getattr(exc, "winerror", None) is not None:
            parts.append(f"winerror={exc.winerror}")
    if isinstance(exc, ssl.SSLError) and getattr(exc, "reason", None):
        parts.append(f"ssl_reason={exc.reason}")
    cause = exc.__cause__
    if cause is not None:
        parts.append(f"cause={type(cause).__name__}: {cause}")
    return " | ".join(parts)


def _smtp_failure_hint(exc: BaseException, *, smtp_server: str = "", port: int = 0) -> str:
    host_hint = f" ({smtp_server!r}:{port})" if smtp_server else ""
    if isinstance(exc, socket.gaierror):
        return f"DNS resolution failed for SMTP host{host_hint} — check SMTP_SERVER/SMTP_HOST."
    if isinstance(exc, (TimeoutError, socket.timeout)):
        return "SMTP connection or operation timed out — check firewall, port, and SMTP_TIMEOUT."
    if isinstance(exc, ConnectionRefusedError):
        return "SMTP connection refused — verify host, port, and that the server accepts inbound SMTP."
    if isinstance(exc, ssl.SSLError):
        return "TLS/SSL handshake failed — try SMTP_SSL=1 on port 465, SMTP_USE_TLS=1 on 587, or SMTP_SSL_VERIFY=0 on corporate proxies."
    if isinstance(exc, smtplib.SMTPAuthenticationError):
        return "SMTP authentication failed — verify SMTP_USER and SMTP_PASS (Gmail requires an App Password)."
    if isinstance(exc, smtplib.SMTPServerDisconnected):
        return "SMTP server closed the connection — often wrong port/TLS mode, idle timeout, or oversized attachments."
    if isinstance(exc, smtplib.SMTPNotSupportedError):
        return "SMTP server does not support the requested TLS mode — check SMTP_SSL / SMTP_USE_TLS and port."
    if isinstance(exc, OSError) and getattr(exc, "errno", None) in (11001,):
        return "SMTP host could not be resolved or reached (Windows WSAHOST_NOT_FOUND)."
    if isinstance(exc, OSError) and getattr(exc, "errno", None) in (101, 113, 111):
        return "SMTP network unreachable or connection refused — check host, port, and agent egress/firewall."
    return "See exception details above."


def _smtp_close(server: smtplib.SMTP | smtplib.SMTP_SSL | None) -> None:
    if server is None:
        return
    try:
        server.quit()
    except Exception as quit_exc:
        logger.debug("[gcp-email] SMTP quit failed: %s", quit_exc)
        try:
            server.close()
        except Exception:
            pass


def _smtp_send_once(
    msg: EmailMessage,
    *,
    smtp_server: str,
    port: int,
    smtp_user: str,
    smtp_pass: str,
) -> None:
    timeout_raw = getenv_any("SMTP_TIMEOUT", default="120")
    try:
        timeout = max(10, int(timeout_raw))
    except ValueError:
        timeout = 120

    use_ssl = _truthy_env("SMTP_SSL") or port == 465
    use_tls = (not use_ssl) and _truthy_env("SMTP_USE_TLS", default="1")
    context = _smtp_ssl_context()

    logger.info(
        "[gcp-email] SMTP connect host=%r port=%s ssl=%s starttls=%s timeout=%ss user=%r",
        smtp_server,
        port,
        use_ssl,
        use_tls,
        timeout,
        smtp_user,
    )

    server: smtplib.SMTP | smtplib.SMTP_SSL | None = None
    try:
        if use_ssl:
            server = smtplib.SMTP_SSL(
                smtp_server,
                port,
                timeout=timeout,
                context=context,
            )
        else:
            server = smtplib.SMTP(timeout=timeout)
            code, resp = server.connect(smtp_server, port)
            logger.info("[gcp-email] SMTP connect response: %s %s", code, resp.decode(errors="replace") if isinstance(resp, bytes) else resp)

        ehlo_code, ehlo_resp = server.ehlo()
        logger.debug(
            "[gcp-email] SMTP EHLO: %s %s",
            ehlo_code,
            ehlo_resp.decode(errors="replace")[:200] if isinstance(ehlo_resp, bytes) else str(ehlo_resp)[:200],
        )

        if use_tls:
            if not server.has_extn("starttls"):
                raise smtplib.SMTPNotSupportedError("STARTTLS not advertised by SMTP server")
            server.starttls(context=context)
            server.ehlo()

        server.login(smtp_user, smtp_pass)
        refused = server.send_message(msg)
        if refused:
            raise smtplib.SMTPRecipientsRefused(refused)
    finally:
        _smtp_close(server)


def _send_smtp_with_retry(
    msg: EmailMessage,
    *,
    smtp_server: str,
    port: int,
    smtp_user: str,
    smtp_pass: str,
    receiver: str,
    max_attempts: int = 3,
    delay_seconds: int = 10,
) -> bool:
    last_exc: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            _smtp_send_once(
                msg,
                smtp_server=smtp_server,
                port=port,
                smtp_user=smtp_user,
                smtp_pass=smtp_pass,
            )
            logger.info(
                "[gcp-email] Email sent to %s (attempt %d/%d)",
                receiver,
                attempt,
                max_attempts,
            )
            return True
        except Exception as exc:
            last_exc = exc
            detail = _format_smtp_error(exc)
            hint = _smtp_failure_hint(exc, smtp_server=smtp_server, port=port)
            logger.warning(
                "[gcp-email] SMTP attempt %d/%d failed: %s",
                attempt,
                max_attempts,
                detail,
            )
            logger.warning("[gcp-email] SMTP diagnosis: %s", hint)
            if attempt < max_attempts:
                logger.info(
                    "[gcp-email] Retrying SMTP in %ss with a fresh connection...",
                    delay_seconds,
                )
                time.sleep(delay_seconds)
    if last_exc is not None:
        logger.error(
            "[gcp-email] Email failed after %d attempts: %s",
            max_attempts,
            _format_smtp_error(last_exc),
        )
        logger.error(
            "[gcp-email] SMTP diagnosis: %s",
            _smtp_failure_hint(last_exc, smtp_server=smtp_server, port=port),
        )
    return False


def _failed_tests_summary_html(rows: list[dict], *, artifact_url: str | None = None) -> str:
    if not rows:
        return (
            '<p class="sub" style="margin:12px 0 16px; font-weight:600; color:#1b5e20;">'
            "No failed tests detected."
            "</p>"
        )
    trs = [
        "<tr>"
        "<th>Suite</th><th>Flow</th><th>Device</th><th>Status</th>"
        "<th>Failure Reason</th><th>AI Analysis</th><th>Screenshot</th><th>Video</th>"
        "</tr>"
    ]
    for row in rows:
        suite = str(row.get("suite") or "—")
        name = str(row.get("test_name") or row.get("flow") or "—")
        device = str(row.get("device") or row.get("device_id") or "—")
        status = str(row.get("status") or "FAIL")
        reason = str(
            row.get("failure_reason") or row.get("reason") or row.get("error_message") or "—"
        )
        ai = str(row.get("ai_analyses") or row.get("ai_analysis") or "—")
        shot = str(row.get("screenshot_artifact") or "").strip()
        video = str(row.get("video_artifact") or "").strip()
        shot_cell = html.escape(shot) if shot else "—"
        video_cell = html.escape(video) if video else "—"
        if artifact_url and shot:
            shot_url = _jenkins_artifact_url(f"build-summary/failed-artifacts/{shot}")
            if shot_url:
                shot_cell = f'<a href="{html.escape(shot_url)}">{html.escape(shot)}</a>'
        if artifact_url and video:
            video_url = _jenkins_artifact_url(f"build-summary/failed-artifacts/{video}")
            if video_url:
                video_cell = f'<a href="{html.escape(video_url)}">{html.escape(video)}</a>'
        cls = _status_html_class(status)
        trs.append(
            "<tr>"
            f'<td class="c-suite">{html.escape(suite)}</td>'
            f'<td class="c-flow">{html.escape(name)}</td>'
            f'<td class="c-dev">{html.escape(device)}</td>'
            f'<td class="{cls}"><strong>{html.escape(status)}</strong></td>'
            f'<td>{html.escape(reason)}</td>'
            f'<td class="c-ai">{html.escape(ai[:200] + ("…" if len(ai) > 200 else ""))}</td>'
            f"<td>{shot_cell}</td>"
            f"<td>{video_cell}</td>"
            "</tr>"
        )
    block = (
        '<p class="sub" style="margin:12px 0 6px; font-weight:600; color:#1f4e79;">'
        "Failed Tests</p>"
        '<table class="ex" role="presentation" style="margin-bottom:18px;">'
        f'{"".join(trs)}</table>'
    )
    if artifact_url:
        block += (
            '<p class="sub"><b>Failed Test Artifacts:</b> '
            f'<a href="{html.escape(artifact_url)}">{html.escape(artifact_url)}</a></p>'
        )
    return block


def _failed_tests_summary_plain(rows: list[dict], *, artifact_url: str | None = None) -> str:
    lines = [
        "Failed Tests",
        "Suite | Flow | Device | Status | Failure Reason | Screenshot | Video",
        "-" * 100,
    ]
    if not rows:
        lines.append("No failed tests detected.")
        if artifact_url:
            lines.append(f"Failed Test Artifacts: {artifact_url}")
        return "\n".join(lines)
    for row in rows:
        suite = str(row.get("suite") or "—")
        name = str(row.get("test_name") or row.get("flow") or "—")
        device = str(row.get("device") or row.get("device_id") or "—")
        status = str(row.get("status") or "FAIL")
        reason = str(row.get("failure_reason") or row.get("reason") or "—")
        shot = str(row.get("screenshot_artifact") or "").strip() or "—"
        video = str(row.get("video_artifact") or "").strip() or "—"
        lines.append(f"{suite} | {name} | {device} | {status} | {reason} | {shot} | {video}")
    if artifact_url:
        lines.append(f"Failed Test Artifacts: {artifact_url}")
    return "\n".join(lines)


def resolve_ai_intelligence_artifacts(root: Path) -> list[Path]:
    """
    intelligent_platform outputs (when present):
    - ai_intelligence_report.xlsx (AI Analyses workbook)
    - intelligence_result.json (full pipeline result)
    """
    r = root.resolve()
    out: list[Path] = []
    seen: set[Path] = set()

    def _add(p: Path) -> None:
        if not p.is_file():
            return
        key = p.resolve()
        if key in seen:
            return
        seen.add(key)
        out.append(p)

    for env_name in ("ORCH_AI_INTELLIGENCE_XLSX", "AI_INTELLIGENCE_REPORT_XLSX"):
        raw = os.getenv(env_name, "").strip()
        if not raw:
            continue
        p = Path(raw)
        if not p.is_absolute():
            p = (r / p).resolve()
        if p.is_file():
            _add(p)
            break

    if not any(p.suffix.lower() == ".xlsx" for p in out):
        _add(r / "build-summary" / "ai_intelligence_report.xlsx")
    if not any(p.suffix.lower() == ".xlsx" for p in out):
        alts = [p for p in r.rglob("ai_intelligence_report.xlsx") if p.is_file()]
        if alts:
            _add(min(alts, key=lambda p: (len(p.parts), str(p))))

    _add(r / "build-summary" / "intelligence_result.json")
    return out


def _normalize_header(s: str) -> str:
    return str(s or "").strip().lower().replace("  ", " ")


def _row_values_to_len(row, n_cols: int) -> list:
    r = list(row) if row is not None else []
    if len(r) < n_cols:
        r.extend([None] * (n_cols - len(r)))
    return r


def _sheet_headers_from_ws(ws) -> list[str] | None:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row1:
        return None
    return [str(h or "").strip() for h in row1]


def _sheet_looks_flow_tabular(headers: list[str]) -> bool:
    hlow = { _normalize_header(x) for x in headers if str(x or "").strip() }
    if "status" not in hlow and "test status" not in hlow:
        return False
    if "suite" not in hlow and "test suite" not in hlow and "suite name" not in hlow:
        return False
    return any(
        k in hlow
        for k in (
            "flow",
            "flow name",
            "test name",
        )
    )


def _ordered_tabular_candidate_names(wb) -> list[str]:
    """Sheets that look like execution data, ordered: Flow Report, Raw Results, then others."""
    out: list[str] = []
    for want in ("Flow Report", "Raw Results"):
        if want in wb.sheetnames and want not in out:
            out.append(want)
    for w in wb.sheetnames:
        wstr = str(w or "").strip()
        if wstr.lower() in ("flow report", "raw results"):
            if w not in out:
                out.append(w)
    for w in wb.sheetnames:
        if w in out:
            continue
        h = _sheet_headers_from_ws(wb[w])
        if h and _sheet_looks_flow_tabular(h):
            out.append(w)
    return out


def _col_index(headers: list[str], *candidates: str) -> int | None:
    lower = { _normalize_header(h): i for i, h in enumerate(headers) if str(h or "").strip()}
    for c in candidates:
        key = _normalize_header(c)
        if key in lower:
            return lower[key]
    return None


def _ai_indices_in_order(headers: list[str]) -> list[int]:
    out: list[int] = []
    for cand in (
        "AI Analysis",
        "AI Analyses",
        "AI Failure Summary",
        "AI Status",
        "Root Cause",
        "Suggested Fix",
    ):
        i = _col_index(headers, cand)
        if i is not None and i not in out:
            out.append(i)
    return out


def _ai_cell_to_email(raw: str) -> str:
    t = (raw or "").strip()
    if not t or t in ("N/A", "NOT_CHECKED"):
        return "—"
    return t


def _flow_cell_invalid(flow_raw: str) -> bool:
    t = (flow_raw or "").strip()
    if not t:
        return True
    if t in (
        "—",
        "–",
        "―",
        "-",
        "—",  # unicode
        "n/a",
        "N/A",
    ):
        return True
    return False


def _simplified_resolve_device(disp: str, did: str) -> str:
    """Display-only: map stored serials to friendly names for email tables."""
    return render_device_display(disp, did)


def _apply_display_to_email_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Final render pass: ensure device column shows friendly names, not serials."""
    out: list[dict[str, str]] = []
    for r in rows:
        nr = dict(r)
        nr["device"] = render_device_display(nr.get("device", ""), nr.get("device_id", ""))
        out.append(nr)
    return out


def _git_branch_for_summary(sheet_kv: dict[str, str]) -> str:
    branch = (sheet_kv.get("Git Branch") or "").strip()
    if branch and branch.lower() != "unknown":
        return branch
    return detect_git_branch(_REPO)


def _parse_table_rows_for_sheet(
    headers: list[str], rows_iter, _sheet_title: str = ""
) -> list[dict[str, str]]:
    n_cols = len(headers)

    i_suite = _col_index(
        headers, "Suite", "Test suite", "Suite Name"
    )
    i_flow = _col_index(
        headers,
        "Flow",
        "Flow Name",
        "Test Name",
    )
    i_dname = _col_index(
        headers,
        "Device Name",
        "Device",
    )
    i_did = _col_index(
        headers,
        "Device ID",
        "Device Id",
        "Udid",
        "UDID",
        "Serial",
    )
    i_status = _col_index(headers, "Status", "Test status")
    i_exit = _col_index(headers, "Exit Code", "Exit code", "ExitCode", "exit_code")
    i_reason = _col_index(
        headers,
        "Error Message",
        "Failure Step",
        "Failure Reason",
        "Reason",
    )
    ai_idx_list = _ai_indices_in_order(headers)

    if i_status is None or i_flow is None:
        return []

    out: list[dict[str, str]] = []
    for row in rows_iter:
        if not row and n_cols:
            continue
        cells = _row_values_to_len(row, n_cols)
        if all(v is None or str(v).strip() == "" for v in cells):
            continue

        def _cell(i: int | None) -> str:
            if i is None or i < 0 or i >= len(cells):
                return ""
            v = cells[i]
            return "" if v is None else str(v).strip()

        suite = _cell(i_suite) if i_suite is not None else ""
        flow = _cell(i_flow)
        if _flow_cell_invalid(flow):
            continue

        st = (_cell(i_status) or "").upper() or "UNKNOWN"
        ex = _cell(i_exit)
        if ex == "":
            ex = "0"

        disp = _cell(i_dname) if i_dname is not None else ""
        did = _cell(i_did) if i_did is not None else ""
        device = _simplified_resolve_device(disp, did)

        raw_ai_parts: list[str] = []
        for j in ai_idx_list:
            t = _cell(j)
            if t and t not in raw_ai_parts:
                raw_ai_parts.append(t)
        raw_ai = " | ".join(raw_ai_parts) if raw_ai_parts else ""
        ai = _ai_cell_to_email(raw_ai)
        reason = _cell(i_reason) if i_reason is not None else ""
        if not reason:
            reason = "—"

        out.append(
            {
                "suite": suite,
                "flow": flow,
                "device": device,
                "device_id": did,
                "status": st,
                "exit_code": ex,
                "ai_analyses": ai,
                "failure_reason": reason,
            }
        )
    return out


def _drop_unknown_mixed_simpler(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """If any row is PASS/FAIL/FLAKY, drop only rows that are only UNKNOWN and look like spurious rollups."""
    if not rows:
        return rows
    has_pfb = any((r.get("status") or "").upper() in ("PASS", "FAIL", "FLAKY") for r in rows)
    if not has_pfb:
        return rows
    return [r for r in rows if (r.get("status") or "").upper() != "UNKNOWN"]


def read_execution_table_rows(
    excel_path: Path,
) -> tuple[list[dict[str, str]], str | None]:
    """
    Return flow-wise rows: suite, flow, device, status, exit_code, ai_analyses.
    Picks the best data sheet and maps columns robustly. Does not use read_only so
    short rows are padded to the header width (fixes ragged/merged cell reads).
    """
    path = excel_path.resolve()
    _log = logger.info
    _print = print
    _log("[email] final_execution Excel path: %s", path)
    _print(f"[orch.mail] final_execution Excel path: {path}")

    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        names = _ordered_tabular_candidate_names(wb)
        _print(f"[orch.mail] workbook sheets: {list(wb.sheetnames)}")
        _log("Available sheet names: %s", wb.sheetnames)
        if not names:
            return [], "No tabular execution sheet found (need Suite, Flow, Status in header row 1)."

        parsed_triples: list[tuple[str, list[dict[str, str]], list[str]]] = []
        for name in names:
            ws = wb[name]
            h = _sheet_headers_from_ws(ws)
            if not h or not _sheet_looks_flow_tabular(h):
                _log("Skip sheet (header mismatch): %s", name)
                continue
            pr = _parse_table_rows_for_sheet(
                h, ws.iter_rows(min_row=2, values_only=True), name
            )
            n = len(pr)
            _log("Sheet %r: parsed %d row(s) before project-wide filters", name, n)
            _print(
                f"[orch.mail] sheet {name!r}: {n} row(s) parsed; header: {h!r}"
            )
            parsed_triples.append((name, pr, h))

        if not parsed_triples:
            return [], "No execution data rows (after parsing candidate sheets)."

        pref = {"Flow Report": 0, "Raw Results": 1}
        sheet_used, out, headers_used = max(
            parsed_triples,
            key=lambda t: (len(t[1]), -pref.get(t[0], 3)),
        )
        out = _drop_unknown_mixed_simpler(out)
        if not out:
            return (
                [],
                f"All rows on sheet {sheet_used!r} were filtered (empty/invalid flow or UNKNOWN-only).",
            )

        _print(
            f"[orch.mail] selected sheet: {sheet_used!r} | total email rows: {len(out)}"
        )
        _print(f"[orch.mail] detected column headers: {headers_used!r}")
        _log("Selected sheet: %r; detected columns: %s", sheet_used, headers_used)
        _log("Using %d email table row(s)", len(out))
        return out, None
    finally:
        wb.close()


def read_summary_sheet_key_values(excel_path: Path) -> dict[str, str]:
    """Key/value pairs from the 'Summary' sheet (column A = label, B = value), like the Excel preview."""
    out: dict[str, str] = {}
    path = excel_path.resolve()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Summary" not in wb.sheetnames:
            return out
        ws = wb["Summary"]
        for r in range(1, 20):
            k = ws.cell(r, 1).value
            v = ws.cell(r, 2).value
            if k is None or str(k).strip() == "":
                continue
            key = str(k).strip()
            if not key:
                continue
            val = "" if v is None else str(v).strip()
            if r == 1 and not val and "kodak" in key.lower() and "smile" in key.lower():
                continue
            out[key] = val
    finally:
        wb.close()
    return out


def compute_summary_from_rows(table_rows: list[dict[str, str]]) -> dict[str, str]:
    """When there is no Summary sheet, match merged-report semantics from Raw rows."""
    t = len(table_rows)
    passed = 0
    flaky = 0
    for r in table_rows:
        st = (r.get("status") or "").upper()
        if st == "PASS":
            passed += 1
        elif st == "FLAKY":
            flaky += 1
    non_pass = t - passed
    return {
        "Total rows": str(t),
        "Passed": str(passed),
        "Flaky": str(flaky),
        "Failed (non-PASS)": str(non_pass - flaky),
    }


def build_summary_display_pairs(
    sheet_kv: dict[str, str], table_rows: list[dict[str, str]], generated_on: str
) -> list[tuple[str, str]]:
    """
    Build ordered (label, value) lines for the email Build Summary section.
    """
    build_no = getenv_any("BUILD_NUMBER", "BUILD_ID", default="").strip() or "—"
    branch = _git_branch_for_summary(sheet_kv)
    exec_time = (
        sheet_kv.get("Generated", "").strip()
        or sheet_kv.get("Generated on", "").strip()
        or generated_on
    )

    if sheet_kv:
        total_s = sheet_kv.get("Total rows") or sheet_kv.get("Total") or "0"
        passed_s = sheet_kv.get("Passed") or "0"
        failed_s = (
            sheet_kv.get("Failed (non-PASS, excl. flaky count below)")
            or sheet_kv.get("Failed (non-PASS)")
            or sheet_kv.get("Failed")
            or "0"
        )
    else:
        comp = compute_summary_from_rows(table_rows)
        total_s = comp.get("Total rows", "0")
        passed_s = comp.get("Passed", "0")
        failed_s = comp.get("Failed (non-PASS)", "0")

    try:
        total_i = int(str(total_s).strip() or "0")
        passed_i = int(str(passed_s).strip() or "0")
    except ValueError:
        total_i = len(table_rows)
        passed_i = sum(1 for r in table_rows if (r.get("status") or "").upper() == "PASS")

    return [
        ("Build", f"#{build_no}"),
        ("Branch", branch),
        ("Execution Time", exec_time),
        ("Total Tests", str(total_i)),
        ("Passed", str(passed_i)),
        ("Failed", str(max(0, total_i - passed_i))),
        ("Pass Rate", _compute_pass_rate(passed_i, total_i)),
    ]


def _summary_stats_html(pairs: list[tuple[str, str]]) -> str:
    if not pairs:
        return ""
    trs = []
    for label, value in pairs:
        trs.append(
            "<tr>"
            f'<th scope="row" style="text-align:left; padding:6px 10px; border:1px solid #ccc; background:#e8f0f8; font-weight:600; white-space:nowrap;">{html.escape(label)}</th>'
            f'<td style="padding:6px 10px; border:1px solid #ccc;">{html.escape(value)}</td>'
            "</tr>"
        )
    return (
        '<p class="sub" style="margin:12px 0 6px; font-weight:600; color:#1f4e79;">'
        "Build Summary</p>"
        f'<table class="sum" role="presentation" style="border-collapse:collapse; max-width:480px; margin-bottom:18px;">{"".join(trs)}</table>'
    )


def _attachments_block_html(attachment_labels: list[str]) -> str:
    if not attachment_labels:
        return ""
    items = "\n    ".join(f"<li>{html.escape(f)}</li>" for f in attachment_labels)
    return f"""<p class="sub"><b>Attachments</b></p>
  <ul style="margin:8px 0 16px; padding-left:20px; color:#333;">
    {items}
  </ul>"""


def _format_ai_for_html(ai_text: str) -> str:
    """Shorter in-cell, full text in title for long AI summaries."""
    t = (ai_text or "").strip() or "—"
    one_line = " ".join(t.split())
    if len(t) > 200:
        short = t[:200] + "…"
        return (
            f'<td class="c-ai" title="{html.escape(one_line, quote=True)}">'
            f"{html.escape(short)}</td>"
        )
    return f'<td class="c-ai">{html.escape(t)}</td>'


def _status_html_class(status: str) -> str:
    u = (status or "").upper()
    if u == "PASS":
        return "st-pass"
    if u in ("FAIL", "PARSE_ERROR", "ERROR"):
        return "st-fail"
    if u == "FLAKY":
        return "st-flaky"
    return "st-other"


def build_email_html(
    rows: list[dict[str, str]],
    generated_on: str,
    error_note: str | None,
    attachment_labels: list[str] | None = None,
    summary_pairs: list[tuple[str, str]] | None = None,
    failed_summary_rows: list[dict] | None = None,
    failed_summary_enabled: bool = False,
    *,
    artifact_url: str | None = None,
    warnings: list[str] | None = None,
    download_links: list[tuple[str, str]] | None = None,
) -> str:
    if failed_summary_enabled:
        table_body = _failed_tests_summary_html(
            failed_summary_rows or rows,
            artifact_url=artifact_url,
        )
        if error_note:
            table_body = f'<p class="warn">{html.escape(error_note)}</p>{table_body}'
    elif error_note and not rows:
        table_body = (
            f'<p class="warn">{html.escape(error_note)}</p>'
            "<p><em>No failed tests to display.</em></p>"
        )
    else:
        trs = []
        for r in rows:
            cls = _status_html_class(r["status"])
            trs.append(
                "<tr>"
                f'<td class="c-suite">{html.escape(r.get("suite", ""))}</td>'
                f'<td class="c-flow">{html.escape(r.get("flow", ""))}</td>'
                f'<td class="c-dev">{html.escape(r.get("device", ""))}</td>'
                f'<td class="{cls}"><strong>{html.escape(r.get("status", ""))}</strong></td>'
                f'<td>{html.escape(r.get("failure_reason", "—"))}</td>'
                f'{_format_ai_for_html(r.get("ai_analyses", ""))}'
                "</tr>"
            )
        if not trs:
            table_body = (
                f'<p class="warn">{html.escape(error_note or "No failed tests in report.")}</p>'
            )
        else:
            thead = (
                "<tr><th>Suite</th><th>Flow</th><th>Device</th><th>Status</th>"
                "<th>Failure Reason</th><th>AI Analysis</th></tr>"
            )
            prefix = f'<p class="warn">{html.escape(error_note)}</p>' if error_note else ""
            table_body = f'{prefix}<table class="ex">{thead}{"".join(trs)}</table>'

    warn_html = ""
    if warnings:
        items = "".join(f"<li>{html.escape(w)}</li>" for w in warnings)
        warn_html = f'<ul class="warn" style="list-style:disc;padding-left:20px;">{items}</ul>'

    links_html = ""
    if download_links:
        items = "".join(
            f'<li><a href="{html.escape(url)}">{html.escape(label)}</a></li>'
            for label, url in download_links
        )
        links_html = (
            '<p class="sub"><b>Download links</b></p>'
            f'<ul style="margin:8px 0 16px; padding-left:20px;">{items}</ul>'
        )

    return f"""\
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: Calibri, "Segoe UI", Arial, sans-serif; font-size: 14px; color: #1a1a1a; }}
  h1 {{ color: #1f4e79; font-size: 20px; margin-bottom: 8px; }}
  .sub {{ color: #666; font-size: 13px; margin-bottom: 16px; }}
  .warn {{ color: #a94442; background: #fbe8e6; padding: 8px; border-radius: 4px; }}
  table.ex {{ border-collapse: collapse; width: 100%; max-width: 1000px; border: 1px solid #000; }}
  table.ex th, table.ex td {{ border: 1px solid #000; padding: 8px 10px; text-align: left; vertical-align: top; word-break: break-word; }}
  .c-ai {{ max-width: 360px; font-size: 13px; line-height: 1.35; color: #222; }}
  table.ex th {{ background: #2e5c8a; color: #fff; font-weight: 600; }}
  .st-pass {{ color: #1b5e20; background: #e8f5e9; font-weight: bold; }}
  .st-fail {{ color: #b71c1c; background: #ffebee; font-weight: bold; }}
  .st-flaky {{ color: #e65100; background: #fff3e0; font-weight: bold; }}
  .st-other {{ color: #333; background: #f5f5f5; font-weight: bold; }}
</style>
</head>
<body>
  <h1>{html.escape(EXECUTION_SUMMARY_TITLE)}</h1>
  {_summary_stats_html(summary_pairs if summary_pairs else [("Generated on", generated_on)])}
  {warn_html}
  {table_body}
  {links_html}
  {_attachments_block_html(attachment_labels or [])}
  <p class="sub" style="margin-top:20px;">This message was sent by Jenkins automation.</p>
</body>
</html>"""


def build_email_plain(
    rows: list[dict[str, str]],
    generated_on: str,
    error_note: str | None,
    attachment_labels: list[str] | None = None,
    summary_pairs: list[tuple[str, str]] | None = None,
    failed_summary_rows: list[dict] | None = None,
    failed_summary_enabled: bool = False,
    *,
    artifact_url: str | None = None,
    warnings: list[str] | None = None,
    download_links: list[tuple[str, str]] | None = None,
) -> str:
    lines = [
        EXECUTION_SUMMARY_TITLE,
        "",
        "## Build Summary",
        "",
    ]
    for label, value in summary_pairs or [("Generated on", generated_on)]:
        lines.append(f"{label}: {value}")
    lines.append("")
    if warnings:
        lines.append("Warnings:")
        for w in warnings:
            lines.append(f"  - {w}")
        lines.append("")
    if failed_summary_enabled or rows:
        lines.append(_failed_tests_summary_plain(failed_summary_rows or rows, artifact_url=artifact_url))
        lines.append("")
    if error_note:
        lines.append(error_note)
        lines.append("")
    if download_links:
        lines.append("Download links:")
        for label, url in download_links:
            lines.append(f"  - {label}: {url}")
        lines.append("")
    lines.append("Attachments:")
    for name in attachment_labels or []:
        lines.append(f"  - {name}")
    lines.append("")
    lines.append(
        "Failed-test logs, screenshots, and videos are in failed_tests_artifacts.zip when attached, "
        "or via Jenkins artifact links when the zip exceeds Gmail size limits."
    )
    return "\n".join(lines)


def getenv_any(*names: str, default: str = "") -> str:
    for name in names:
        v = os.getenv(name, "").strip()
        if v:
            return v
    return default


def _orch_email_attach_ai() -> bool:
    """Set ORCH_EMAIL_ATTACH_AI=1 to add ai_intelligence_report.xlsx + intelligence_result.json."""
    return getenv_any("ORCH_EMAIL_ATTACH_AI", default="").lower() in ("1", "true", "yes", "on")


def _orch_email_strict() -> bool:
    return (os.environ.get("ORCH_EMAIL_STRICT", "").strip().lower() in ("1", "true", "yes", "on"))


def _smtp_config_ready() -> bool:
    """
    True when user/pass/receiver and server (or implied Gmail) are all present to attempt SMTP.
    """
    smtp_user = getenv_any("SMTP_USER", "SENDER_EMAIL", "GMAIL_USER")
    smtp_server = getenv_any("SMTP_SERVER", "SMTP_HOST")
    if not smtp_server and smtp_user and "@" in smtp_user and smtp_user.lower().rstrip().endswith(
        ("@gmail.com", "@googlemail.com")
    ):
        smtp_server = "smtp.gmail.com"
    smtp_pass = getenv_any("SMTP_PASS", "SMTP_PASSWORD", "SENDER_PASSWORD", "GMAIL_APP_PASSWORD")
    receiver = getenv_any("RECEIVER_EMAIL", "MAIL_TO", "EMAIL_RECIPIENTS", "RECIPIENT", "ORCH_MAIL_TO")
    return bool(smtp_server and smtp_user and smtp_pass and receiver)


def send_execution_report_email(
    excel_path: Path,
    *,
    root: Path | None = None,
    subject: str | None = None,
    body: str | None = None,
) -> bool:
    """
    Returns True if mail was sent, False if skipped or failed (logged).
    HTML body is built from final_execution_report.xlsx unless ``body`` is set (overrides).
    """
    excel_path = excel_path.resolve()
    if not excel_path.is_file():
        logger.error("Excel attachment missing: %s", excel_path)
        return False

    smtp_user = getenv_any("SMTP_USER", "SENDER_EMAIL", "GMAIL_USER")
    smtp_server = getenv_any("SMTP_SERVER", "SMTP_HOST")
    if not smtp_server and smtp_user and "@" in smtp_user and smtp_user.lower().rstrip().endswith(
        ("@gmail.com", "@googlemail.com")
    ):
        smtp_server = "smtp.gmail.com"
        logger.info("Defaulting SMTP server to smtp.gmail.com (Gmail sender)")

    smtp_port_raw = getenv_any("SMTP_PORT", default="587")
    smtp_pass = getenv_any(
        "SMTP_PASS",
        "SMTP_PASSWORD",
        "SENDER_PASSWORD",
        "GMAIL_APP_PASSWORD",
    )
    sender = getenv_any("SENDER_EMAIL", "SMTP_USER", "GMAIL_USER")
    receiver = getenv_any(
        "RECEIVER_EMAIL",
        "MAIL_TO",
        "EMAIL_RECIPIENTS",
        "RECIPIENT",
        "ORCH_MAIL_TO",
    )

    gen_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rroot = root.resolve() if root is not None else None
    exc_resolved = excel_path.resolve()

    failed_summary_rows: list[dict] = []
    failed_summary_enabled = False
    failed_zip: Path | None = None
    artifact_url: str | None = None
    zip_size = 0
    videos_count = 0
    screenshots_count = 0

    if rroot is not None:
        failed_summary_rows, failed_summary_enabled = load_failed_tests_summary(rroot)
        failed_zip = resolve_failed_tests_artifacts_zip(rroot)
        if failed_zip is not None:
            zip_size = _file_size(failed_zip)
            artifact_url = _jenkins_artifact_url("build-summary/failed_tests_artifacts.zip")
        videos_count = sum(1 for r in failed_summary_rows if r.get("video_artifact"))
        screenshots_count = sum(1 for r in failed_summary_rows if r.get("screenshot_artifact"))

    table_rows, table_err = [], None
    if body is None:
        table_rows, table_err = read_execution_table_rows(excel_path)
        table_rows = _apply_display_to_email_rows(table_rows)

    all_rows_for_stats = table_rows
    failed_email_rows = _filter_failed_email_rows(table_rows)
    failed_email_rows = _enrich_rows_from_failed_summary(failed_email_rows, failed_summary_rows)

    if failed_summary_enabled and failed_summary_rows:
        seen = {
            (
                str(r.get("suite") or "").casefold(),
                str(r.get("flow") or "").strip(),
                str(r.get("device_id") or "").strip(),
            )
            for r in failed_email_rows
        }
        for item in failed_summary_rows:
            key = (
                str(item.get("suite") or "").casefold(),
                str(item.get("test_name") or item.get("flow") or "").strip(),
                str(item.get("device_id") or "").strip(),
            )
            if key in seen:
                continue
            seen.add(key)
            failed_email_rows.append(
                {
                    "suite": str(item.get("suite") or ""),
                    "flow": str(item.get("test_name") or item.get("flow") or ""),
                    "device": render_device_display("", str(item.get("device_id") or "")),
                    "device_id": str(item.get("device_id") or ""),
                    "status": str(item.get("status") or "FAIL"),
                    "failure_reason": str(item.get("failure_reason") or "—"),
                    "ai_analyses": "—",
                    "video_artifact": str(item.get("video_artifact") or ""),
                    "screenshot_artifact": str(item.get("screenshot_artifact") or ""),
                }
            )

    sheet_kv = read_summary_sheet_key_values(excel_path) if body is None else {}
    summary_pairs = (
        build_summary_display_pairs(sheet_kv, all_rows_for_stats, gen_ts) if body is None else []
    )

    failed_count = len(failed_email_rows)
    total_count = len(all_rows_for_stats)
    subj = subject or getenv_any("ORCH_MAIL_SUBJECT", default="").strip()
    if not subj:
        subj = _build_email_subject(failed_count=failed_count, total_count=total_count)

    ai_candidates: list[tuple[Path, str]] = []
    if rroot is not None and _orch_email_attach_ai():
        for ap in resolve_ai_intelligence_artifacts(rroot):
            if ap.resolve() != exc_resolved:
                ai_candidates.append((ap, ap.name))

    attach_candidates: list[tuple[Path, str]] = [(excel_path, excel_path.name)]
    attach_candidates.extend(ai_candidates)

    force_link_names: set[str] = set()
    if failed_zip is not None and zip_size > MAX_ATTACHMENT_SIZE:
        force_link_names.add(failed_zip.name)
        logger.warning(
            "[gcp-email] ZIP Size: %s exceeds %s — zip will use download link",
            _format_bytes(zip_size),
            _format_bytes(MAX_ATTACHMENT_SIZE),
        )
    elif failed_zip is not None and failed_summary_enabled:
        attach_candidates.append(
            (failed_zip, f"{failed_zip.name} (failed test logs, screenshots, videos)")
        )

    plan = _plan_attachments(
        attach_candidates,
        force_link_names=frozenset(force_link_names),
    )
    attachment_labels = list(plan.attach_labels)
    warnings = list(plan.warnings)
    download_links = list(plan.download_links)

    if plan.link_only_mode and artifact_url and failed_zip is not None:
        warnings.append(
            f"failed_tests_artifacts.zip ({_format_bytes(zip_size)}) available via Jenkins artifact link."
        )

    print(f"[gcp-email] ZIP Size: {_format_bytes(zip_size)}", flush=True)
    print(f"[gcp-email] Total Attachment Size: {_format_bytes(plan.total_bytes)}", flush=True)
    print(f"[gcp-email] Failed Tests Count: {failed_count}", flush=True)
    print(f"[gcp-email] Videos Attached: {videos_count}", flush=True)
    print(f"[gcp-email] Screenshots Attached: {screenshots_count}", flush=True)
    print(f"[gcp-email] Artifact URL: {artifact_url or '—'}", flush=True)

    if body is not None:
        text_body = body
        html_body = f"<html><body><pre>{html.escape(body)}</pre></body></html>"
    else:
        text_body = build_email_plain(
            failed_email_rows,
            gen_ts,
            table_err,
            attachment_labels,
            summary_pairs,
            failed_summary_rows,
            failed_summary_enabled or bool(failed_email_rows),
            artifact_url=artifact_url,
            warnings=warnings,
            download_links=download_links,
        )
        html_body = build_email_html(
            failed_email_rows,
            gen_ts,
            table_err,
            attachment_labels,
            summary_pairs,
            failed_summary_rows,
            failed_summary_enabled or bool(failed_email_rows),
            artifact_url=artifact_url,
            warnings=warnings,
            download_links=download_links,
        )

    if not smtp_server or not smtp_user or not smtp_pass or not receiver:
        logger.warning(
            "Email skipped: set at minimum SMTP_USER (or SENDER_EMAIL), "
            "SMTP_PASS (or SMTP_PASSWORD), RECEIVER_EMAIL (or MAIL_TO), and "
            "SMTP_SERVER (or omit for @gmail.com to default to smtp.gmail.com). "
            "Gmail: use an App Password, not the normal account password."
        )
        return False

    try:
        port = int(smtp_port_raw)
    except ValueError:
        port = 587

    msg = EmailMessage()
    msg["Subject"] = subj
    msg["From"] = sender or smtp_user
    msg["To"] = receiver
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")

    for ap in plan.attach_paths:
        _add_file_attachment(msg, ap)

    return _send_smtp_with_retry(
        msg,
        smtp_server=smtp_server,
        port=port,
        smtp_user=smtp_user,
        smtp_pass=smtp_pass,
        receiver=receiver,
    )


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    root = Path(os.environ.get("WORKSPACE", os.getcwd())).resolve()
    excel = resolve_final_excel_path(root)
    if excel is None:
        return 1
    if not _smtp_config_ready():
        if _orch_email_strict():
            logger.error(
                "ORCH_EMAIL_STRICT=1: SMTP is not fully configured. "
                "Jenkins: in the Send Final Email stage set SMTP_USER, SMTP_PASS, RECEIVER_EMAIL (e.g. in the batch step), "
                "or export them on the agent before: python mailout/send_email.py"
            )
            return 1
        logger.warning(
            "SMTP not configured (set SMTP_USER + SMTP_PASS + RECEIVER_EMAIL, and SMTP_SERVER or @gmail.com); "
            "exiting 0. Set ORCH_EMAIL_STRICT=1 to fail this step if mail is required."
        )
        return 0
    return 0 if send_execution_report_email(excel, root=root) else 1


if __name__ == "__main__":
    raise SystemExit(main())
