"""Generate AI_Agent_Report.xlsx — always written, pass or fail."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from executor.status_parser import REPORT_COLUMNS


def write_excel_report(
    path: Path,
    rows: list[dict[str, Any]],
    *,
    session_metrics: Any = None,
    module_summaries: list[Any] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    all_columns = list(REPORT_COLUMNS)
    normalized = [_normalize_row(r, all_columns) for r in rows] if rows else [_blank_row()]

    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font  # type: ignore

        wb = Workbook()
        ws = wb.active
        ws.title = "AI Agent Runs"
        ws.append(all_columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in normalized:
            ws.append([row.get(c, "") for c in all_columns])

        summary = wb.create_sheet("Execution Summary")
        passed = sum(1 for r in normalized if str(r.get("Status", "")).upper() == "PASS")
        failed = sum(1 for r in normalized if str(r.get("Status", "")).upper() == "FAIL")
        recovered = sum(
            1 for r in normalized if str(r.get("Recovery Result", "")).upper() == "PASS"
        )
        summary.append(["Metric", "Value"])
        summary.append(["Total Test Rows", len(normalized)])
        summary.append(["Passed", passed])
        summary.append(["Failed", failed])
        summary.append(["Recovered Failures", recovered])
        if session_metrics:
            sm = session_metrics
            summary.append(["Application Launch Count", getattr(sm, "launch_count", 0)])
            summary.append(["Application Restart Count", getattr(sm, "restart_count", 0)])
            summary.append(
                ["Adaptive Wait Savings (sec)", round(getattr(sm, "adaptive_wait_saved_sec", 0), 1)]
            )
            summary.append(["Navigation Optimisations", getattr(sm, "navigation_optimisations", 0)])
            summary.append(["Checkpoint Resumes", getattr(sm, "checkpoint_resumes", 0)])
            summary.append(["Artifacts Learned", getattr(sm, "artifacts_learned", 0)])
            summary.append(
                ["Knowledge Base Updated", "Yes" if getattr(sm, "knowledge_updated", False) else "No"]
            )
            summary.append(
                ["Avg Screen Load (sec)", round(getattr(sm, "avg_screen_load_sec", 0), 2)]
            )
            prev = getattr(sm, "performance_vs_previous_pct", 0)
            if prev:
                summary.append(["Performance vs Previous Run", f"{prev:+.1f}%"])
        health_vals = [
            float(str(r.get("Overall Health Score", "0")).rstrip("%") or 0)
            for r in normalized
            if r.get("Overall Health Score")
        ]
        if health_vals:
            summary.append(["Overall Health Score", f"{health_vals[-1]:.0f}%"])

        if module_summaries:
            mod_sheet = wb.create_sheet("Module Journey")
            mod_sheet.append(["Module", "Status", "Duration (sec)"])
            for m in module_summaries:
                mod_sheet.append([m.name, m.status, round(m.duration_sec, 1)])

        wb.save(path)
        print(f"[ai-agent.report] wrote {path}", flush=True)
    except ImportError:
        _write_csv_fallback(path.with_suffix(".csv"), normalized, all_columns)


def _normalize_row(row: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    out = {c: row.get(c, "") for c in columns}
    if not out.get("AI Decision"):
        out["AI Decision"] = "No AI Intervention"
    return out


def _blank_row() -> dict[str, Any]:
    row = {c: "" for c in REPORT_COLUMNS}
    row["AI Decision"] = "No AI Intervention"
    row["Status"] = "UNKNOWN"
    return row


def _write_csv_fallback(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    import csv

    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
